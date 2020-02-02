#!/usr/bin/env python3

import openpyxl
import argparse
from openpyxl import Workbook

"""
Script to compare two xlsx file, highlighting common sequences.


Copyright 2020 Margherita Maria Ferrari.


This file is part of SequencesComparator.

SequencesComparator is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

SequencesComparator is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with SequencesComparator.  If not, see <http://www.gnu.org/licenses/>.
"""


class SequencesComparator:
    @classmethod
    def get_args(cls):
        parser = argparse.ArgumentParser(description='Sequences comparator')
        parser.add_argument('-i1', '--file1', metavar='IN_FILE1', type=str, required=True,
                            help='First XLSX input file', default=None)
        parser.add_argument('-i2', '--file2', metavar='IN_FILE2', type=str, required=True,
                            help='Second XLSX input file', default=None)
        parser.add_argument('-m', '--min-freq', metavar='MIN_FREQ', type=int, required=False,
                            help='Minimum frequency for a sequence to be considered valid', default=1)
        parser.add_argument('-mc', '--min-freq-common', metavar='MIN_FREQ_COMMON', type=int, required=False,
                            help='Minimum frequency for a common sequence to be considered valid', default=1)
        parser.add_argument('-o', '--output-file', metavar='OUTPUT_FILE', type=str, required=False,
                            help='Output XLSX file', default='output.xlsx')
        return parser.parse_args()

    @classmethod
    def compare_files(cls, in_file1=None, in_file2=None, min_freq=1, min_freq_common=1):
        if not in_file1 or not in_file2:
            raise AssertionError('Not enough files specified')

        # File 1
        wb = openpyxl.load_workbook(in_file1, read_only=True)  # workbook - open file
        ws = wb.active  # worksheet - the active worksheet is the first sheet
        data1 = [x for x in list(ws.rows) if x[1].value >= min_freq]
        wb.close()

        # File 2
        wb = openpyxl.load_workbook(in_file2, read_only=True)
        ws = wb.active
        data2 = [x for x in list(ws.rows) if int(x[1].value) >= min_freq]
        wb.close()

        data_common = list()
        data_common_bad = list()
        for row1 in list(data1):
            for row2 in [x for x in data2 if str(x[0].value) == str(row1[0].value)]:
                if (int(row1[1].value) >= min_freq_common) or (int(row2[1].value) >= min_freq_common):
                    data_common.append((str(row1[0].value), int(row1[1].value), int(row2[1].value)))
                else:
                    data_common_bad.append((str(row1[0].value), int(row1[1].value), int(row2[1].value)))
                data1.remove(row1)
                data2.remove(row2)

        ret = {'common': data_common,
               'common_bad': data_common_bad,
               'only_first': data1,
               'only_second': data2
               }

        return ret

    @classmethod
    def save_results(cls, data=None, output_file='output.xlsx'):
        if not data:
            raise AssertionError('No data to save')
        if not output_file:
            output_file = 'output.xlsx'

        wb = Workbook(write_only=True)

        # Common sequences
        ws = wb.create_sheet('Common Seq')
        for row in data.get('common', list()):
            ws.append((row[0], str(row[1]) + ' -- ' + str(row[2])))
        ws.close()

        # Common sequences with low frequency
        ws = wb.create_sheet('Common Seq Low Freq')
        for row in data.get('common_bad', list()):
            ws.append((row[0], str(row[1]) + ' -- ' + str(row[2])))
        ws.close()

        # Sequences only in file 1
        ws = wb.create_sheet('Only File 1')
        for row in data.get('only_first', list()):
            ws.append((str(row[0].value), str(row[1].value)))
        ws.close()

        # Sequences only in file 1
        ws = wb.create_sheet('Only File 2')
        for row in data.get('only_second', list()):
            ws.append((str(row[0].value), str(row[1].value)))
        ws.close()

        wb.save(output_file)


if __name__ == '__main__':
    args = vars(SequencesComparator.get_args())
    min_freq_count = args.get('min_freq')
    min_freq_common_count = args.get('min_freq_common')

    if not min_freq_count or min_freq_count < 0:
        min_freq_count = 1

    if not min_freq_common_count or min_freq_common_count < 0:
        min_freq_common_count = 1

    results = SequencesComparator.compare_files(args.get('file1'), args.get('file2'),
                                                min_freq_count, min_freq_common_count)
    SequencesComparator.save_results(results, args.get('output_file'))
